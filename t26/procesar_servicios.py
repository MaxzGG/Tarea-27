#Maximiliano Gomez Guevara
#2177546
import subprocess
import csv
import openpyxl
from openpyxl.utils import get_column_letter

# Función para ejecutar el script de PowerShell
def ejecutar_powershell(script_path):
    try:
        # Ejecutar el script de PowerShell
        result = subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-File", script_path], check=True)
        return result.returncode
    except subprocess.CalledProcessError as e:
        print(f"Error al ejecutar PowerShell: {e}")
        return None

# Función para convertir CSV a Excel
def csv_a_excel(csv_path, excel_path):
    try:
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Servicios"

        # Abrir el archivo CSV y leer su contenido
        with open(csv_path, mode='r', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Guardar el archivo Excel
        wb.save(excel_path)
        print(f"Datos guardados en {excel_path}")

    except Exception as e:
        print(f"Error al convertir CSV a Excel: {e}")

# Ruta de los archivos
ruta_powershell = r"D:\Programacion Clase\t27\monitor_servicios.ps1"
ruta_csv = r"D:\Programacion Clase\t27\monitor_servicios.ps1"
ruta_excel = r"D:\Programacion Clase\t27\monitor_servicios.ps1"

# Ejecutar el script de PowerShell y procesar la salida
if ejecutar_powershell(ruta_powershell) == 0:
    csv_a_excel(ruta_csv, ruta_excel)
else:
    print("No se pudo ejecutar el script de PowerShell.")